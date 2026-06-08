#!/usr/bin/env python3
"""
Convert the three ORB result CSVs into Excel workbooks with live formulas
for all calculated columns (Trade_PL, Drawdown_Pct, Cumulative_PL).
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT_W = Font(bold=True, size=11, color="FFFFFF")
WIN_FILL = PatternFill(start_color="E2EFDA", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FCE4EC", fill_type="solid")
NOOP_FILL = PatternFill(start_color="F5F5F5", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style='thin', color='D9D9D9'),
)

NUM_2DP = '#,##0.00'
NUM_PCT = '0.00"%"'
NUM_INT = '#,##0'


def build_workbook(csv_path, xlsx_path, is_intraday=True):
    """
    Read the CSV and produce an Excel workbook with formulas.

    Column layout (intraday):
      A: Date, B: Day_of_Week, C: Symbol, D: Range_High, E: Range_Low,
      F: Range (formula), G: Trade_Direction, H: Entry_Price, I: Exit_Price,
      J: Trade_PL (formula), K: Drawdown_Pct, L: Result, M: Cumulative_PL (formula)

    Column layout (monthly/quarterly):
      A: Period, B: Symbol, C: Range_Days, D: Trade_Days, E: Range_High,
      F: Range_Low, G: Range (formula), H: Trade_Direction, I: Entry_Price,
      J: Exit_Price, K: Trade_PL (formula), L: Drawdown_Pct, M: Result,
      N: Cumulative_PL (formula)
    """
    rows = []
    with open(csv_path, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "ORB Results"

    if is_intraday:
        headers = [
            'Date', 'Day_of_Week', 'Symbol', 'Range_High', 'Range_Low',
            'Range', 'Trade_Direction', 'Entry_Price', 'Exit_Price',
            'Trade_PL', 'Drawdown_Pct', 'Result', 'Cumulative_PL',
        ]
        col_rh, col_rl, col_range = 'D', 'E', 'F'
        col_dir, col_entry, col_exit = 'G', 'H', 'I'
        col_pl, col_dd, col_result, col_cum = 'J', 'K', 'L', 'M'
    else:
        headers = [
            'Period', 'Symbol', 'Range_Days', 'Trade_Days', 'Range_High',
            'Range_Low', 'Range', 'Trade_Direction', 'Entry_Price',
            'Exit_Price', 'Trade_PL', 'Drawdown_Pct', 'Result',
            'Cumulative_PL',
        ]
        col_rh, col_rl, col_range = 'E', 'F', 'G'
        col_dir, col_entry, col_exit = 'H', 'I', 'J'
        col_pl, col_dd, col_result, col_cum = 'K', 'L', 'M', 'N'

    # --- Header row ---
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT_W
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # --- Data rows ---
    for ri, row in enumerate(rows, 2):
        if is_intraday:
            ws[f'A{ri}'] = row['Date']
            ws[f'B{ri}'] = row['Day_of_Week']
            ws[f'C{ri}'] = row['Symbol']
        else:
            ws[f'A{ri}'] = row['Period']
            ws[f'B{ri}'] = row['Symbol']
            ws[f'C{ri}'] = _num(row.get('Range_Days', ''))
            ws[f'D{ri}'] = _num(row.get('Trade_Days', ''))

        ws[f'{col_rh}{ri}'] = _num(row['Range_High'])
        ws[f'{col_rl}{ri}'] = _num(row['Range_Low'])

        # Range = Range_High - Range_Low (formula)
        ws[f'{col_range}{ri}'] = f'={col_rh}{ri}-{col_rl}{ri}'

        ws[f'{col_dir}{ri}'] = row['Trade_Direction']
        ws[f'{col_entry}{ri}'] = _num(row['Entry_Price'])
        ws[f'{col_exit}{ri}'] = _num(row['Exit_Price'])

        # Trade P/L formula:
        # IF direction="Long", Exit-Entry; IF "Short", Entry-Exit; else 0
        ws[f'{col_pl}{ri}'] = (
            f'=IF({col_dir}{ri}="Long",{col_exit}{ri}-{col_entry}{ri},'
            f'IF({col_dir}{ri}="Short",{col_entry}{ri}-{col_exit}{ri},0))'
        )

        # Drawdown % (raw value — this comes from the simulation, not formulaic)
        ws[f'{col_dd}{ri}'] = _num(row['Drawdown_Pct'])

        ws[f'{col_result}{ri}'] = row['Result']

        # Cumulative P/L formula: previous cumulative + this row's PL
        if ri == 2:
            ws[f'{col_cum}{ri}'] = f'={col_pl}{ri}'
        else:
            ws[f'{col_cum}{ri}'] = f'={col_cum}{ri-1}+{col_pl}{ri}'

        # Conditional row fill
        result = row['Result']
        fill = None
        if result == 'Win':
            fill = WIN_FILL
        elif result == 'Loss':
            fill = LOSS_FILL
        elif result == 'No-Op':
            fill = NOOP_FILL

        if fill:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = fill

        for ci in range(1, len(headers) + 1):
            ws.cell(row=ri, column=ci).border = THIN_BORDER

    # --- Number formatting ---
    last_row = len(rows) + 1
    for col in [col_rh, col_rl, col_range, col_entry, col_exit, col_pl, col_cum]:
        for ri in range(2, last_row + 1):
            ws[f'{col}{ri}'].number_format = NUM_2DP
    for ri in range(2, last_row + 1):
        ws[f'{col_dd}{ri}'].number_format = '0.00'

    # --- Summary sheet ---
    ws_sum = wb.create_sheet("Summary")
    ws_sum['A1'] = 'Metric'
    ws_sum['B1'] = 'Value'
    ws_sum['A1'].font = HEADER_FONT_W
    ws_sum['B1'].font = HEADER_FONT_W
    ws_sum['A1'].fill = HEADER_FILL
    ws_sum['B1'].fill = HEADER_FILL

    data_range = f"'ORB Results'!{col_result}2:{col_result}{last_row}"
    pl_range = f"'ORB Results'!{col_pl}2:{col_pl}{last_row}"
    dd_range = f"'ORB Results'!{col_dd}2:{col_dd}{last_row}"

    summary_rows = [
        ('Total Rows', f'=COUNTA({data_range})'),
        ('Trades (excl No-Op)', f'=COUNTIF({data_range},"Win")+COUNTIF({data_range},"Loss")+COUNTIF({data_range},"EOD-Close")+COUNTIF({data_range},"Period-Close")'),
        ('Wins', f'=COUNTIF({data_range},"Win")'),
        ('Losses', f'=COUNTIF({data_range},"Loss")'),
        ('EOD/Period Close', f'=COUNTIF({data_range},"EOD-Close")+COUNTIF({data_range},"Period-Close")'),
        ('No-Op', f'=COUNTIF({data_range},"No-Op")'),
        ('Win Rate (W/L only)', f'=IF(B4+B5>0,B4/(B4+B5),0)'),
        ('Win Rate (all trades)', f'=IF(B3>0,COUNTIFS({data_range},"<>No-Op",{pl_range},">"&0)/B3,0)'),
        ('Total P/L', f'=SUM({pl_range})'),
        ('Avg Trade P/L', f'=IF(B3>0,B10/B3,0)'),
        ('Avg Drawdown %', f'=AVERAGEIFS({dd_range},{data_range},"<>No-Op")'),
        ('Avg Win Drawdown %', f'=AVERAGEIF({data_range},"Win",{dd_range})'),
        ('Final Cumulative P/L', f"='ORB Results'!{col_cum}{last_row}"),
    ]

    for i, (label, formula) in enumerate(summary_rows, 2):
        ws_sum[f'A{i}'] = label
        ws_sum[f'B{i}'] = formula

    ws_sum['B8'].number_format = '0.00%'
    ws_sum['B9'].number_format = '0.00%'
    ws_sum['B10'].number_format = NUM_2DP
    ws_sum['B11'].number_format = NUM_2DP
    ws_sum['B12'].number_format = '0.00'
    ws_sum['B13'].number_format = '0.00'
    ws_sum['B14'].number_format = NUM_2DP

    # Column widths
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 15
    ws_sum.column_dimensions['A'].width = 25
    ws_sum.column_dimensions['B'].width = 18

    # Freeze top row
    ws.freeze_panes = 'A2'

    wb.save(xlsx_path)
    print(f"  Saved {xlsx_path} ({last_row - 1} data rows)")


def _num(val):
    if val is None or val == '' or val == 'None':
        return None
    try:
        f = float(val)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return val


def main():
    print("Converting to Excel with formulas ...")

    build_workbook(
        'mnq_orb_results.csv',
        'mnq_orb_results.xlsx',
        is_intraday=True,
    )

    build_workbook(
        'mnq_monthly_orb.csv',
        'mnq_monthly_orb.xlsx',
        is_intraday=False,
    )

    build_workbook(
        'mnq_quarterly_orb.csv',
        'mnq_quarterly_orb.xlsx',
        is_intraday=False,
    )

    build_workbook(
        'mnq_yearly_orb.csv',
        'mnq_yearly_orb.xlsx',
        is_intraday=False,
    )

    print("Done.")


if __name__ == '__main__':
    main()
