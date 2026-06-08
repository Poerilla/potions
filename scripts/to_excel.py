#!/usr/bin/env python3
"""
v2 Excel writer — converts the pre-placed OCO stop backtest CSVs into
formatted Excel workbooks with live formulas, summary sheet, and conditional
fills. Supports MNQ and MYM (1-min granularity, v2 stop-entry model).

The archived v1 version of this script lives at
`../archived/v1_scripts/to_excel.py` and reads the old v1 CSVs.
"""
import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter


HEADER_FONT_W = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
WIN_FILL      = PatternFill(start_color="E2EFDA", fill_type="solid")
LOSS_FILL     = PatternFill(start_color="FCE4EC", fill_type="solid")
EOD_WIN_FILL  = PatternFill(start_color="D9E7F5", fill_type="solid")
EOD_LOSS_FILL = PatternFill(start_color="FBE5D6", fill_type="solid")
THIN_BORDER   = Border(bottom=Side(style='thin', color='D9D9D9'))

NUM_2DP = '#,##0.00'

DEFAULTS = {
    'MNQ': {
        'src': '/home/tester/hsm/potions/mnq/mnq_orb_results_stops.csv',
        'dst': '/home/tester/hsm/potions/mnq/mnq_orb_results_stops.xlsx',
        'mult': 2.0,
    },
    'MYM': {
        'src': '/home/tester/hsm/potions/mym/mym_orb_results_stops.csv',
        'dst': '/home/tester/hsm/potions/mym/mym_orb_results_stops.xlsx',
        'mult': 0.5,
    },
}


def _num(val):
    if val in (None, '', 'None'):
        return None
    try:
        f = float(val)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return val


def build_workbook(csv_path, xlsx_path, mult):
    rows = []
    with open(csv_path, 'r') as f:
        for row in csv.DictReader(f):
            rows.append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "v2 ORB Results"

    # Column layout
    headers = [
        'Date', 'Day_of_Week', 'Symbol',
        'Range_High', 'Range_Low', 'Range',
        'Trade_Direction', 'Entry_Price', 'Exit_Price',
        'Trade_PL_pts', 'Net_$', 'Result',
        'Cumulative_pts', 'Cumulative_$',
    ]
    cols = {h: chr(ord('A') + i) for i, h in enumerate(headers)}

    # Header row
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT_W
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')

    # Data rows
    for ri, row in enumerate(rows, 2):
        ws[f'A{ri}'] = row['Date']
        ws[f'B{ri}'] = row['Day_of_Week']
        ws[f'C{ri}'] = row['Symbol']
        ws[f'D{ri}'] = _num(row['Range_High'])
        ws[f'E{ri}'] = _num(row['Range_Low'])
        ws[f'F{ri}'] = f'=D{ri}-E{ri}'
        ws[f'G{ri}'] = row['Trade_Direction']
        ws[f'H{ri}'] = _num(row['Entry_Price'])
        ws[f'I{ri}'] = _num(row['Exit_Price'])
        ws[f'J{ri}'] = (
            f'=IF(G{ri}="Long",I{ri}-H{ri},'
            f'IF(G{ri}="Short",H{ri}-I{ri},0))'
        )
        ws[f'K{ri}'] = f'=J{ri}*{mult}-1.5'    # net after $1.50 RT fee
        ws[f'L{ri}'] = row['Result']
        ws[f'M{ri}'] = f'=J{ri}' if ri == 2 else f'=M{ri-1}+J{ri}'
        ws[f'N{ri}'] = f'=K{ri}' if ri == 2 else f'=N{ri-1}+K{ri}'

        fill = None
        result = row['Result']
        if result == 'Win':
            fill = WIN_FILL
        elif result == 'Loss':
            fill = LOSS_FILL
        elif result == 'EOD-Win':
            fill = EOD_WIN_FILL
        elif result == 'EOD-Loss':
            fill = EOD_LOSS_FILL

        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            if fill:
                cell.fill = fill
            cell.border = THIN_BORDER

    last_row = len(rows) + 1

    # Number formats
    for col in ('D', 'E', 'F', 'H', 'I', 'J', 'K', 'M', 'N'):
        for ri in range(2, last_row + 1):
            ws[f'{col}{ri}'].number_format = NUM_2DP

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum['A1'] = 'Metric'; ws_sum['B1'] = 'Value'
    for c in ('A1', 'B1'):
        ws_sum[c].font = HEADER_FONT_W
        ws_sum[c].fill = HEADER_FILL

    result_rng = f"'v2 ORB Results'!L2:L{last_row}"
    pl_rng     = f"'v2 ORB Results'!J2:J{last_row}"
    net_rng    = f"'v2 ORB Results'!K2:K{last_row}"

    summary = [
        ('Total Trades',      f'=COUNTA({result_rng})'),
        ('Wins (incl EOD-Win)', f'=COUNTIF({result_rng},"Win")+COUNTIF({result_rng},"EOD-Win")'),
        ('Losses (incl EOD-Loss)', f'=COUNTIF({result_rng},"Loss")+COUNTIF({result_rng},"EOD-Loss")'),
        ('Win Rate',          f'=B3/B2'),
        ('Total P/L (pts)',   f'=SUM({pl_rng})'),
        ('Total Net $',       f'=SUM({net_rng})'),
        ('Avg Trade Net $',   f'=AVERAGE({net_rng})'),
        ('Max Single Win $',  f'=MAX({net_rng})'),
        ('Max Single Loss $', f'=MIN({net_rng})'),
        ('Final Cum Net $',   f"='v2 ORB Results'!N{last_row}"),
    ]
    for i, (label, formula) in enumerate(summary, 2):
        ws_sum[f'A{i}'] = label
        ws_sum[f'B{i}'] = formula

    ws_sum['B5'].number_format = '0.00%'
    for r in (6, 7, 8, 9, 10, 11):
        ws_sum[f'B{r}'].number_format = NUM_2DP

    # Column widths
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 15
    ws_sum.column_dimensions['A'].width = 28
    ws_sum.column_dimensions['B'].width = 16

    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    print(f"  Saved {xlsx_path}  ({last_row - 1} trades)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--product', choices=list(DEFAULTS.keys()), default='all')
    ap.add_argument('--src', help='Override input CSV')
    ap.add_argument('--dst', help='Override output XLSX')
    ap.add_argument('--mult', type=float, help='$ per point (e.g., 2.0 for MNQ)')
    args = ap.parse_args()

    print("Writing v2 ORB Excel workbook(s) ...")
    if args.src or args.dst or args.mult:
        src = args.src or DEFAULTS['MNQ']['src']
        dst = args.dst or DEFAULTS['MNQ']['dst']
        mult = args.mult if args.mult is not None else DEFAULTS['MNQ']['mult']
        build_workbook(src, dst, mult)
    else:
        products = DEFAULTS.keys() if args.product == 'all' else [args.product]
        for p in products:
            cfg = DEFAULTS[p]
            if not Path(cfg['src']).exists():
                print(f"  Skipping {p} — {cfg['src']} not found")
                continue
            build_workbook(cfg['src'], cfg['dst'], cfg['mult'])
    print("Done.")


if __name__ == '__main__':
    main()
